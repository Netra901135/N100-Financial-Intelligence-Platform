# Updated engine.py (Day 15)
import sqlite3
import yaml
import pandas as pd
from pathlib import Path
from openpyxl.styles import PatternFill
from openpyxl import load_workbook

class ScreenerEngine:

    def __init__(self):
        self.db = Path("database/nifty100.db")
        self.conn = sqlite3.connect(self.db)

        self.df = pd.read_sql("SELECT * FROM financial_ratios", self.conn)
        self.sectors = pd.read_sql("SELECT * FROM sectors", self.conn)

        self.df = self.df.merge(
            self.sectors,
            on="company_id",
            how="left",
        )

        with open("config/screener_config.yaml", "r") as f:
            self.config = yaml.safe_load(f)

    def load_preset(self, preset):
        return self.config[preset]
    
    def winsorize(self, series):
        """
        Cap values at the 10th and 90th percentile.
        """
        s = pd.to_numeric(series, errors="coerce")
        p10 = s.quantile(0.10)
        p90 = s.quantile(0.90)
        return s.clip(lower=p10, upper=p90)
    
    def normalize_score(self, series, higher_is_better=True):
        """
        Convert a metric to a 0-100 score.
        """
        s = self.winsorize(series)
        minimum = s.min()
        maximum = s.max()
        if pd.isna(minimum) or pd.isna(maximum) or minimum == maximum:
            return pd.Series(50, index=s.index)
        
        score = (s - minimum) / (maximum - minimum) * 100
        if not higher_is_better:
            score = 100 - score
        return score.fillna(0)
    
    def sector_normalize(self, df, column, higher_is_better=True):
        """
        Normalize a metric within each broad sector.
        """
        return (
            df.groupby("broad_sector")[column]
            .transform(
                lambda x: self.normalize_score(
                    x,
                    higher_is_better=higher_is_better,
                )
            )
        )
    def apply_filter(self, preset):
        rules = self.load_preset(preset)
        df = self.df.copy()

        df = (
            df.sort_values("year")
            .groupby("company_id", as_index=False)
            .tail(1)
        )

        if "roe_min" in rules:
            df = df[df["return_on_equity_pct"] >= rules["roe_min"]]

        if "roce_min" in rules:
            df = df[df["return_on_capital_employed_pct"] >= rules["roce_min"]]

        if "npm_min" in rules:
            df = df[df["net_profit_margin_pct"] >= rules["npm_min"]]

        if "opm_min" in rules:
            df = df[df["operating_profit_margin_pct"] >= rules["opm_min"]]

        if "debt_to_equity_max" in rules:
            df = df[
                (df["broad_sector"] == "Financials") |
                (df["debt_to_equity"] <= rules["debt_to_equity_max"])
            ]

        if "interest_coverage_min" in rules:
            icr = df["interest_coverage"].copy()
            icr = icr.replace("Debt Free", float("inf"))
            icr = pd.to_numeric(icr, errors="coerce").fillna(float("inf"))
            df = df[icr >= rules["interest_coverage_min"]]

        if "asset_turnover_min" in rules:
            df = df[df["asset_turnover"] >= rules["asset_turnover_min"]]

        if "free_cash_flow_min" in rules:
            df = df[df["free_cash_flow"] >= rules["free_cash_flow_min"]]

        if "dividend_payout_min" in rules:
            df = df[df["dividend_payout"] >= rules["dividend_payout_min"]]

        if "dividend_payout_max" in rules:
            df = df[df["dividend_payout"] <= rules["dividend_payout_max"]]

        if "revenue_cagr_5yr_min" in rules and "revenue_cagr_5yr" in df.columns:
            df = df[df["revenue_cagr_5yr"] >= rules["revenue_cagr_5yr_min"]]

        if "pat_cagr_5yr_min" in rules and "pat_cagr_5yr" in df.columns:
            df = df[df["pat_cagr_5yr"] >= rules["pat_cagr_5yr_min"]]

        if "eps_cagr_5yr_min" in rules and "eps_cagr_5yr" in df.columns:
            df = df[df["eps_cagr_5yr"] >= rules["eps_cagr_5yr_min"]]

        if "pe_max" in rules and "pe_ratio" in df.columns:
            df = df[df["pe_ratio"] <= rules["pe_max"]]

        if "pb_max" in rules and "pb_ratio" in df.columns:
            df = df[df["pb_ratio"] <= rules["pb_max"]]

        if "dividend_yield_min" in rules and "dividend_yield" in df.columns:
            df = df[df["dividend_yield"] >= rules["dividend_yield_min"]]

        if "sales_min" in rules and "sales" in df.columns:
            df = df[df["sales"] >= rules["sales_min"]]

        if "market_cap_min" in rules and "market_cap" in df.columns:
            df = df[df["market_cap"] >= rules["market_cap_min"]]

        if "net_profit_min" in rules and "net_profit" in df.columns:
            df = df[df["net_profit"] >= rules["net_profit_min"]]

        if (
            "revenue_cagr_3yr_min" in rules
            and "revenue_cagr_3yr" in df.columns
        ):
            df = df[
                df["revenue_cagr_3yr"] >= rules["revenue_cagr_3yr_min"]
            ]
        if (
            "debt_to_equity_declining" in rules and "debt_to_equity_declining" in df.columns
        ):
            df = df[
                df["debt_to_equity_declining"] == rules["debt_to_equity_declining"]
            ]
        # Profitability
        df["roe_score"] = self.sector_normalize(df, "return_on_equity_pct")
        df["roce_score"] = self.sector_normalize(df, "return_on_capital_employed_pct")
        df["npm_score"] = self.sector_normalize(df, "net_profit_margin_pct")

# Cash Quality
        df["fcf_cagr_score"] = self.sector_normalize(df, "fcf_cagr_5yr")
        df["cfo_pat_score"] = self.sector_normalize(df, "cfo_pat_ratio")
        df["fcf_flag_score"] = (df["free_cash_flow"] > 0).astype(int) * 100

# Growth
        df["revenue_score"] = self.sector_normalize(df, "revenue_cagr_5yr")
        df["pat_score"] = self.sector_normalize(df, "pat_cagr_5yr")

# Leverage
        df["de_score"] = self.sector_normalize(
            df,
            "debt_to_equity",
            higher_is_better=False,
        )
        df["interest_coverage_numeric"] = (
    pd.to_numeric(
        df["interest_coverage"].replace("Debt Free", float("inf")),
        errors="coerce",
    )
)

        df["interest_coverage_numeric"] = (
            df["interest_coverage_numeric"]
            .fillna(df["interest_coverage_numeric"].max())
        )

        df["icr_score"] = self.sector_normalize(
        df,
        "interest_coverage_numeric",
        )

# Composite score
        df["composite_quality_score"] = (
            0.15 * df["roe_score"] +
            0.10 * df["roce_score"] +
            0.10 * df["npm_score"] +
            0.15 * df["fcf_cagr_score"] +
            0.10 * df["cfo_pat_score"] +
            0.05 * df["fcf_flag_score"] +
            0.10 * df["revenue_score"] +
            0.10 * df["pat_score"] +
            0.10 * df["de_score"] +
            0.05 * df["icr_score"]
        ).round(2)
        
        return df.sort_values(
            "composite_quality_score",
            ascending=False,
        )
    
    def export_excel(self):
        from pathlib import Path
        
        output_dir = Path("output")
        output_dir.mkdir(exist_ok=True)

        output_file = output_dir / "screener_output.xlsx"

        with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
            for preset in self.config.keys():
                df = self.apply_filter(preset)
                if df.empty:
                    continue
                
                columns = [
                    "company_id",
                    "broad_sector",
                    "sub_sector",

                    "composite_quality_score",

                    "return_on_equity_pct",
                    "return_on_capital_employed_pct",
                    "net_profit_margin_pct",
                    "operating_profit_margin_pct",

                    "revenue_cagr_5yr",
                    "pat_cagr_5yr",
                    "fcf_cagr_5yr",

                    "cfo_pat_ratio",

                    "free_cash_flow",

                    "debt_to_equity",
                    "interest_coverage",

                    "asset_turnover",

                    "sales",
                    "net_profit",

                    "debt_to_equity_declining",
                    "year",
                ]

                available = [c for c in columns if c in df.columns]

                export_df = (
                    df[available]
                    .sort_values(
                        "composite_quality_score",
                        ascending=False,
                    )
                )

                export_df.to_excel(
                    writer,
                    sheet_name=preset[:31],
                    index=False,
                )

        print(f"\nExcel exported -> {output_file}")

    def colour_excel(self):
        workbook = load_workbook("output/screener_output.xlsx")
        
        green = PatternFill(
            fill_type="solid",
            start_color="C6EFCE",
            end_color="C6EFCE",
        )

        red = PatternFill(
            fill_type="solid",
            start_color="FFC7CE",
            end_color="FFC7CE",
        )

        for sheet in workbook.sheetnames:
            ws = workbook[sheet]
            headers = {}
            for cell in ws[1]:
                headers[cell.column] = cell.value

            rules = self.load_preset(sheet)
            
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    header = headers[cell.column]

                    if header is None:
                        continue

                    if header == "return_on_equity_pct":
                        threshold = rules.get("roe_min")

                        if threshold is not None:

                            if cell.value >= threshold:
                                cell.fill = green
                            else:
                                cell.fill = red

                    elif header == "return_on_capital_employed_pct":

                        threshold = rules.get("roce_min")

                        if threshold is not None:

                            if cell.value >= threshold:
                                cell.fill = green
                            else:
                                cell.fill = red

                    elif header == "net_profit_margin_pct":

                        threshold = rules.get("npm_min")

                        if threshold is not None:

                            if cell.value >= threshold:
                               cell.fill = green
                            else:
                               cell.fill = red

                    elif header == "debt_to_equity":

                        threshold = rules.get("debt_to_equity_max")

                        if threshold is not None:

                            if cell.value <= threshold:
                                cell.fill = green
                            else:
                                cell.fill = red

                        elif header == "interest_coverage":

                            threshold = rules.get("interest_coverage_min")

                            if threshold is not None:

                                value = cell.value

                                if value == "Debt Free":
                                    value = float("inf")

                                if float(value) >= threshold:
                                    cell.fill = green
                                else:
                                    cell.fill = red

        workbook.save("output/screener_output.xlsx")

        print("Conditional formatting applied.")

    def run_all_presets(self):
        for preset in self.config.keys():
            result = self.apply_filter(preset)
            print("\n" + "=" * 70)
            print(f"PRESET : {preset.upper()}")
            print("=" * 70)

            if len(result) == 0:
                print("No companies matched.")
                continue

            print(result[
                [
                    "company_id",
                    "composite_quality_score",
                    "return_on_equity_pct",
                    "return_on_capital_employed_pct",
                    "debt_to_equity",
                    "free_cash_flow",
                ]
            ].head(10))

            print(f"\nTotal Companies : {len(result)}")
        self.export_excel()
        self.colour_excel()

def main():
    engine = ScreenerEngine()
    engine.run_all_presets()


if __name__ == "__main__":
    main()
