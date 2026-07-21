import sqlite3
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

class PeerComparisonReport:

    def __init__(self):

        self.db = Path("database/nifty100.db")
        self.conn = sqlite3.connect(self.db)

        self.df = None
        self.percentiles = None

    def load_data(self):

        # -------------------------------
        # Financial Ratios
        # -------------------------------
        ratios = pd.read_sql(
            """
            SELECT *
            FROM financial_ratios
            """,
            self.conn
        )

        latest_year = ratios["year"].max()

        ratios = ratios[
            ratios["year"] == latest_year
        ]

        # -------------------------------
        # Sector Information
        # -------------------------------
        sectors = pd.read_sql(
            """
            SELECT *
            FROM sectors
            """,
            self.conn
        )

        # -------------------------------
        # Peer Groups
        # -------------------------------
        peer_groups = pd.read_excel(
            "data/raw/peer_groups.xlsx"
        )

        # -------------------------------
        # Percentiles
        # -------------------------------
        percentiles = pd.read_sql(
            """
            SELECT *
            FROM peer_percentiles
            """,
            self.conn
        )

        percentiles = percentiles[
            percentiles["year"] == latest_year
        ]

        # -------------------------------
        # Merge all datasets
        # -------------------------------
        self.df = (
            ratios
            .merge(
                sectors,
                on="company_id",
                how="left"
            )
            .merge(
                peer_groups[
                    [
                        "company_id",
                        "peer_group_name",
                        "is_benchmark"
                    ]
                ],
                on="company_id",
                how="left"
            )
        )

        # Companies not in peer group
        self.df["peer_group_name"] = self.df[
            "peer_group_name"
        ].fillna(
            self.df["broad_sector"]
        )

        self.df["is_benchmark"] = self.df[
            "is_benchmark"
        ].fillna(False)

        self.percentiles = percentiles

        print("Report data loaded.")
        print(f"Latest Year : {latest_year}")
        print(f"Companies   : {len(self.df)}")
        print(
            f"Peer Groups : {self.df['peer_group_name'].nunique()}"
        )

        print("\nPeer Groups\n")
        print(
            sorted(
                self.df["peer_group_name"]
                .dropna()
                .unique()
            )
        )

    def create_excel_report(self):

        report_columns = [
            "company_id",
            "peer_group_name",
            "is_benchmark",
            "sales",
            "return_on_equity_pct",
            "return_on_equity_pct_percentile",
            "return_on_capital_employed_pct",
            "return_on_capital_employed_pct_percentile",
            "dividend_payout",
            "net_profit_margin_pct",
            "net_profit_margin_pct_percentile",
            "operating_profit_margin_pct",
            "return_on_equity_pct",
            "return_on_capital_employed_pct",
            "return_on_assets_pct",
            "debt_to_equity",
            "debt_to_equity_percentile",
            "interest_coverage",
            "interest_coverage_percentile",
            "asset_turnover",
            "asset_turnover_percentile",
            "free_cash_flow",
            "free_cash_flow_percentile",
            "revenue_cagr_5yr",
            "revenue_cagr_5yr_percentile",
            "revenue_cagr_3yr",
            "pat_cagr_5yr",
            "pat_cagr_5yr_percentile",
            "eps_cagr_5yr",
            "eps_cagr_5yr_percentile",
            "cfo_pat_ratio",
            "fcf_cagr_5yr"
        ]

        output_file = "output/peer_comparison.xlsx"

        Path("output").mkdir(exist_ok=True)

        with pd.ExcelWriter(
            output_file,
            engine="openpyxl"
        ) as writer:

            for peer_group in sorted(
                self.df["peer_group_name"].dropna().unique()
            ):

                sheet = self.df[
                    self.df["peer_group_name"] == peer_group
                ][report_columns].copy()
                numeric_columns = [
                    col for col in sheet.columns
                    if pd.api.types.is_numeric_dtype(sheet[col])
                ]

                median_row = {}

                for col in sheet.columns:

                    if col == "company_id":
                        median_row[col] = "Peer Group Median"

                    elif col == "peer_group_name":
                        median_row[col] = peer_group

                    elif col == "is_benchmark":
                        median_row[col] = ""

                    elif col in numeric_columns:
                        median_row[col] = sheet[col].median()

                    else:
                        median_row[col] = ""

               
                sheet.to_excel(
                    writer,
                    sheet_name=peer_group[:31],  # Excel sheet name limit
                    index=False
                )

                print(f"Written sheet: {peer_group}")

        print(f"\nSaved: {output_file}")

    def format_excel_report(self):

        workbook = load_workbook("output/peer_comparison.xlsx")

        green_fill = PatternFill(
            fill_type="solid",
            start_color="C6EFCE"
        )

        yellow_fill = PatternFill(
            fill_type="solid",
            start_color="FFEB9C"
        )

        red_fill = PatternFill(
            fill_type="solid",
            start_color="FFC7CE"
        )

        benchmark_fill = PatternFill(
            fill_type="solid",
            start_color="FFD966"
        )

        for sheet in workbook.worksheets:

            headers = [
                cell.value
                for cell in sheet[1]
            ]

            percentile_cols = []

            for idx, header in enumerate(headers, start=1):

                if (
                    isinstance(header, str)
                    and header.endswith("_percentile")
                ):
                    percentile_cols.append(idx)

            benchmark_col = headers.index(
                "is_benchmark"
            ) + 1

            for row in range(2, sheet.max_row + 1):

                # Highlight benchmark row
                if sheet.cell(
                    row=row,
                    column=benchmark_col
                ).value:

                    for col in range(
                        1,
                        sheet.max_column + 1
                    ):
                        sheet.cell(
                            row=row,
                            column=col
                        ).fill = benchmark_fill

                # Colour percentile cells
                for col in percentile_cols:

                    value = sheet.cell(
                        row=row,
                        column=col
                    ).value

                    if value is None:
                        continue

                    if value >= 75:

                        sheet.cell(
                            row=row,
                            column=col
                        ).fill = green_fill

                    elif value <= 25:

                        sheet.cell(
                            row=row,
                            column=col
                        ).fill = red_fill

                    else:

                        sheet.cell(
                            row=row,
                            column=col
                        ).fill = yellow_fill

            # Header formatting
            for cell in sheet[1]:

                cell.font = Font(bold=True)

        workbook.save(
            "output/peer_comparison.xlsx"
        )

        print("Excel formatting applied.")
    def merge_percentiles(self):
        percentile_df = self.percentiles.pivot_table(
            index="company_id",
            columns="metric",
            values="percentile_rank"
        ).reset_index()

        percentile_df.columns = [
            "company_id"
            if col == "company_id"
            else f"{col}_percentile"
            for col in percentile_df.columns
        ]

        self.df = self.df.merge(
            percentile_df,
            on="company_id",
            how="left"
        )

        print("Percentile columns merged.")
        print(
            percentile_df.head()
        )

    

def main():

    report = PeerComparisonReport()

    report.load_data()
    report.merge_percentiles()
    report.create_excel_report()
    report.format_excel_report()


if __name__ == "__main__":
    main()